"""
Finance API: Financial categories and transactions.
"""
import io
import re
import hashlib

from django.db import transaction
from django.db.models import Q, Sum
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, TruncYear
from django.http import HttpResponse, FileResponse
from django.utils import timezone
from django.conf import settings
from django.core.files.base import ContentFile

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import JSONParser, FormParser, MultiPartParser

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

from .helpers import _is_admin_user, _client_ip, _safe_payload, _create_approval_request, _draw_authenticity_qr
from ..models import FinancialCategory, FinancialTransaction, FinancialDocumentSequence, ReportCertificate
from ..permissions import IsTreasurerOrAdmin
from ..serializers import FinancialCategorySerializer, FinancialTransactionSerializer


class FinancialCategoryViewSet(viewsets.ModelViewSet):
    queryset = FinancialCategory.objects.all().order_by('-id')
    serializer_class = FinancialCategorySerializer
    permission_classes = [IsTreasurerOrAdmin]


class FinancialTransactionViewSet(viewsets.ModelViewSet):
    queryset = FinancialTransaction.objects.all().order_by('-date', '-id')
    serializer_class = FinancialTransactionSerializer
    permission_classes = [IsTreasurerOrAdmin]
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def create(self, request, *args, **kwargs):
        if _is_admin_user(request.user):
            return super().create(request, *args, **kwargs)
        ar = _create_approval_request(request, model='FinancialTransaction', action='create', payload=_safe_payload(request.data))
        return Response({'detail': 'Action soumise à approbation.', 'approval_request_id': ar.id}, status=202)

    def update(self, request, *args, **kwargs):
        if _is_admin_user(request.user):
            return super().update(request, *args, **kwargs)
        tx = self.get_object()
        ar = _create_approval_request(
            request, model='FinancialTransaction', action='update',
            payload=_safe_payload(request.data),
            target_object_id=getattr(tx, 'id', None),
            object_repr=getattr(tx, 'document_number', None) or getattr(tx, 'receipt_code', None) or str(getattr(tx, 'id', '') or ''),
        )
        return Response({'detail': 'Action soumise à approbation.', 'approval_request_id': ar.id}, status=202)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if _is_admin_user(request.user):
            return super().destroy(request, *args, **kwargs)
        tx = self.get_object()
        ar = _create_approval_request(
            request, model='FinancialTransaction', action='delete',
            payload=None,
            target_object_id=getattr(tx, 'id', None),
            object_repr=getattr(tx, 'document_number', None) or getattr(tx, 'receipt_code', None) or str(getattr(tx, 'id', '') or ''),
        )
        return Response({'detail': 'Action soumise à approbation.', 'approval_request_id': ar.id}, status=202)

    def get_queryset(self):
        qs = super().get_queryset()

        event_id = (self.request.query_params.get('event') or '').strip()
        if event_id:
            try:
                qs = qs.filter(event_id=int(event_id))
            except (TypeError, ValueError):
                qs = qs.none()

        direction = (self.request.query_params.get('direction') or '').strip().lower()
        if direction in {'in', 'out'}:
            qs = qs.filter(direction=direction)

        tx_type = (self.request.query_params.get('transaction_type') or '').strip().lower()
        if tx_type:
            qs = qs.filter(transaction_type=tx_type)

        currency = (self.request.query_params.get('currency') or '').strip().upper()
        if currency:
            qs = qs.filter(currency=currency)

        start = (self.request.query_params.get('start') or '').strip()
        end = (self.request.query_params.get('end') or '').strip()
        if start and end:
            qs = qs.filter(date__range=[start, end])
        elif start:
            qs = qs.filter(date__gte=start)
        elif end:
            qs = qs.filter(date__lte=end)

        q = (self.request.query_params.get('q') or '').strip()
        if q:
            qs = qs.filter(
                Q(document_number__icontains=q)
                | Q(receipt_code__icontains=q)
                | Q(reference_number__icontains=q)
                | Q(donor_name__icontains=q)
                | Q(donor_email__icontains=q)
                | Q(recipient_name__icontains=q)
                | Q(recipient_email__icontains=q)
                | Q(recipient_phone__icontains=q)
            )

        return qs

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ModuleNotFoundError:
            return Response(
                {'detail': "Le module 'openpyxl' n'est pas installé sur le serveur. Installez-le puis relancez le backend."},
                status=500,
            )

        wb = Workbook()
        ws = wb.active
        ws.title = 'Transactions'

        headers = [
            'Document', 'Sens', 'Type', 'Montant', 'Devise', 'Date',
            'Donateur', 'Email donateur', 'Bénéficiaire', 'Email bénéficiaire',
            'Téléphone bénéficiaire', 'Mode paiement', 'Référence', 'Description',
            'Caissier', 'Créé par', 'Créé le',
        ]
        ws.append(headers)

        header_fill = PatternFill('solid', fgColor='7C2D12')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='D9D9D9')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.row_dimensions[1].height = 28
        for col_idx, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = header_alignment
            c.border = border

        qs = self.get_queryset().select_related('cashier', 'created_by')
        for tx in qs:
            cashier_name = None
            if tx.cashier:
                cashier_name = (f"{tx.cashier.first_name} {tx.cashier.last_name}").strip() or tx.cashier.username
            created_by_name = None
            if tx.created_by:
                created_by_name = (f"{tx.created_by.first_name} {tx.created_by.last_name}").strip() or tx.created_by.username

            created_at = getattr(tx, 'created_at', None)
            if created_at:
                try:
                    created_at = timezone.localtime(created_at).replace(tzinfo=None)
                except Exception:
                    created_at = None

            ws.append([
                tx.document_number or tx.receipt_code or '',
                tx.direction or '',
                tx.transaction_type or '',
                float(tx.amount) if tx.amount is not None else None,
                (tx.currency or '').upper(),
                getattr(tx, 'date', None),
                tx.donor_name or '',
                tx.donor_email or '',
                tx.recipient_name or '',
                tx.recipient_email or '',
                tx.recipient_phone or '',
                tx.payment_method or '',
                tx.reference_number or '',
                tx.description or '',
                cashier_name or '',
                created_by_name or '',
                created_at,
            ])

        last_row = ws.max_row
        last_col = ws.max_column

        body_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        money_alignment = Alignment(horizontal='right', vertical='top')

        for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
            for cell in row:
                cell.alignment = body_alignment
                cell.border = border

        for r in range(2, last_row + 1):
            ws.cell(row=r, column=4).number_format = '#,##0.00'
            ws.cell(row=r, column=4).alignment = money_alignment
            ws.cell(row=r, column=6).number_format = 'dd/mm/yyyy'
            ws.cell(row=r, column=17).number_format = 'dd/mm/yyyy hh:mm'

        ws.freeze_panes = 'A2'

        if last_row >= 2:
            table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
            table = Table(displayName='FinancialTransactionsTable', ref=table_ref)
            style = TableStyleInfo(
                name='TableStyleMedium9',
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        for col_idx in range(1, last_col + 1):
            max_len = 0
            for cell in ws[get_column_letter(col_idx)]:
                v = cell.value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 40)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        filename = f"transactions_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            out.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'], url_path='time-series')
    def time_series(self, request):
        period = request.query_params.get('period', 'daily').lower()
        start = request.query_params.get('start')
        end = request.query_params.get('end')

        qs = self.get_queryset()

        if start:
            qs = qs.filter(date__gte=start)
        if end:
            qs = qs.filter(date__lte=end)

        if period == 'weekly':
            trunc = TruncWeek('date')
        elif period == 'monthly':
            trunc = TruncMonth('date')
        elif period in {'yearly', 'annual', 'annually'}:
            trunc = TruncYear('date')
        else:
            trunc = TruncDay('date')

        rows = (
            qs.annotate(p=trunc)
            .values('p', 'currency', 'direction')
            .annotate(total=Sum('amount'))
            .order_by('p', 'currency', 'direction')
        )

        series = {}
        for r in rows:
            key = r['p'].date().isoformat() if hasattr(r['p'], 'date') else str(r['p'])
            currency = r['currency']
            direction = r['direction']
            total = float(r['total'] or 0)

            series.setdefault(key, {})
            series[key].setdefault(currency, {'in': 0.0, 'out': 0.0, 'net': 0.0})
            series[key][currency][direction] = total

        for key, curMap in series.items():
            for cur, agg in curMap.items():
                agg['net'] = float(agg.get('in', 0) or 0) - float(agg.get('out', 0) or 0)

        return Response({'period': period, 'start': start, 'end': end, 'series': series})

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        rows = (
            self.get_queryset()
            .values('currency', 'direction')
            .annotate(total=Sum('amount'))
            .order_by('currency', 'direction')
        )

        totals = {}
        for r in rows:
            cur = r['currency'] or 'CDF'
            totals.setdefault(cur, {'in': 0.0, 'out': 0.0, 'net': 0.0})
            direction = r['direction']
            totals[cur][direction] = float(r['total'] or 0)

        for cur, agg in totals.items():
            agg['net'] = float(agg.get('in', 0) or 0) - float(agg.get('out', 0) or 0)

        return Response({'totals': totals})

    @action(detail=False, methods=['get'], url_path='verify-receipt')
    def verify_receipt(self, request):
        code = (request.query_params.get('code') or '').strip()
        if not code:
            return Response({'detail': 'code requis'}, status=400)

        tx = FinancialTransaction.objects.filter(receipt_code=code, direction='in').select_related('cashier', 'created_by').first()
        if not tx:
            return Response({'detail': 'reçu introuvable'}, status=404)

        return Response({
            'id': tx.id,
            'receipt_code': tx.receipt_code,
            'amount': str(tx.amount),
            'currency': tx.currency,
            'date': tx.date.isoformat(),
            'transaction_type': tx.transaction_type,
            'donor_name': tx.donor_name,
            'donor_email': tx.donor_email,
            'cashier_name': (f"{tx.cashier.first_name} {tx.cashier.last_name}").strip() if tx.cashier else None,
            'created_by_name': (f"{tx.created_by.first_name} {tx.created_by.last_name}").strip() if tx.created_by else None,
            'created_at': tx.created_at.isoformat() if tx.created_at else None,
        })

    @action(detail=False, methods=['get'], url_path='verify-document')
    def verify_document(self, request):
        code = (request.query_params.get('code') or '').strip()
        if not code:
            return Response({'detail': 'code requis'}, status=400)

        tx = FinancialTransaction.objects.filter(document_number=code).select_related('cashier', 'created_by').first()
        if not tx:
            tx = FinancialTransaction.objects.filter(receipt_code=code).select_related('cashier', 'created_by').first()
        if not tx:
            return Response({'detail': 'document introuvable'}, status=404)

        return Response({
            'id': tx.id,
            'document_number': tx.document_number,
            'direction': tx.direction,
            'amount': str(tx.amount),
            'currency': tx.currency,
            'date': tx.date.isoformat(),
            'transaction_type': tx.transaction_type,
            'donor_name': tx.donor_name,
            'donor_email': tx.donor_email,
            'recipient_name': getattr(tx, 'recipient_name', None),
            'recipient_email': getattr(tx, 'recipient_email', None),
            'recipient_phone': getattr(tx, 'recipient_phone', None),
            'cashier_name': (f"{tx.cashier.first_name} {tx.cashier.last_name}").strip() if tx.cashier else None,
            'created_by_name': (f"{tx.created_by.first_name} {tx.created_by.last_name}").strip() if tx.created_by else None,
            'created_at': tx.created_at.isoformat() if tx.created_at else None,
        })

    @action(detail=True, methods=['get'], url_path='receipt')
    def receipt(self, request, pk=None):
        tx = self.get_object()
        if tx.direction != 'in':
            return Response({'detail': 'reçu disponible uniquement pour une entrée'}, status=400)

        self._ensure_receipt_pdf(tx, request)
        if not tx.receipt_pdf:
            return Response({'detail': 'impossible de générer le reçu'}, status=500)

        resp = FileResponse(open(tx.receipt_pdf.path, 'rb'), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="receipt_{tx.document_number or tx.receipt_code}.pdf"'
        return resp

    @action(detail=True, methods=['get'], url_path='voucher')
    def voucher(self, request, pk=None):
        tx = self.get_object()
        if tx.direction != 'out':
            return Response({'detail': 'bon disponible uniquement pour une sortie'}, status=400)

        with transaction.atomic():
            self._ensure_document_number(tx)
        pdf = self._build_voucher_pdf_bytes(tx, request)

        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="voucher_{tx.document_number}.pdf"'
        return resp

    @action(detail=False, methods=['get'], url_path='report-pdf')
    def report_pdf(self, request):
        period = request.query_params.get('period', 'daily').lower()
        start = request.query_params.get('start')
        end = request.query_params.get('end')

        qs = FinancialTransaction.objects.all().order_by('date', 'id')
        if start and end:
            qs = qs.filter(date__range=[start, end])
        elif start:
            qs = qs.filter(date__gte=start)
        elif end:
            qs = qs.filter(date__lte=end)

        if period == 'weekly':
            trunc = TruncWeek('date')
        elif period == 'monthly':
            trunc = TruncMonth('date')
        elif period in {'yearly', 'annual', 'annually'}:
            trunc = TruncYear('date')
        else:
            trunc = TruncDay('date')

        rows = (
            qs.annotate(p=trunc)
            .values('p', 'currency', 'direction')
            .annotate(total=Sum('amount'))
            .order_by('p', 'currency', 'direction')
        )

        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)
        meta = self._pdf_brand_layout(
            c,
            doc_title='Rapport',
            doc_subtitle='Document officiel — finance',
            badge_text='RAPPORT',
            badge_rgb=(0.34, 0.32, 0.76),
        )

        width = meta['width']
        height = meta['height']
        margin = meta['margin']
        x = meta['x']
        y = meta['y_start']

        c.setFillColorRGB(0.07, 0.10, 0.16)
        c.setFont('Helvetica-Bold', 18)
        c.drawString(x, y, 'Bordereau financier')

        c.setFillColorRGB(0.40, 0.45, 0.55)
        c.setFont('Helvetica', 10)
        c.drawRightString(width - margin - 10 * mm, y, f"Période: {start or '—'} → {end or '—'}")
        y -= 10 * mm

        c.setFillColorRGB(0.07, 0.10, 0.16)
        c.setFont('Helvetica-Bold', 10)
        period_label = {
            'daily': 'Journalier',
            'weekly': 'Hebdomadaire',
            'monthly': 'Mensuel',
            'yearly': 'Annuel',
            'annual': 'Annuel',
            'annually': 'Annuel',
        }.get(period, period)
        c.drawString(x, y, f"Granularité: {period_label}")
        y -= 10 * mm

        c.setFont('Helvetica-Bold', 10)
        c.setFillColorRGB(0.40, 0.45, 0.55)
        c.drawString(x, y, 'Période')
        c.drawString(x + 55 * mm, y, 'Devise')
        c.drawRightString(x + 112 * mm, y, 'Entrées')
        c.drawRightString(x + 138 * mm, y, 'Sorties')
        c.drawRightString(width - margin - 10 * mm, y, 'Solde')

        y -= 6 * mm
        c.setFont('Helvetica', 10)
        c.setFillColorRGB(0.20, 0.24, 0.30)
        series = {}
        for r in rows:
            key = r['p'].date().isoformat() if hasattr(r['p'], 'date') else str(r['p'])
            cur = r['currency']
            direction = r['direction']
            total = float(r['total'] or 0)
            series.setdefault(key, {})
            series[key].setdefault(cur, {'in': 0.0, 'out': 0.0, 'net': 0.0})
            series[key][cur][direction] = total

        for key, curMap in series.items():
            for cur, agg in curMap.items():
                agg['net'] = float(agg.get('in', 0) or 0) - float(agg.get('out', 0) or 0)

        for p, curMap in series.items():
            for cur, agg in curMap.items():
                if y < 20 * mm:
                    c.showPage()
                    meta_p = self._pdf_brand_layout(
                        c,
                        doc_title='Rapport',
                        doc_subtitle='Document officiel — finance',
                        badge_text='RAPPORT',
                        badge_rgb=(0.34, 0.32, 0.76),
                    )
                    width = meta_p['width']
                    height = meta_p['height']
                    margin = meta_p['margin']
                    x = meta_p['x']
                    y = meta_p['y_start']
                    c.setFont('Helvetica', 10)
                    c.setFillColorRGB(0.20, 0.24, 0.30)

                c.drawString(x, y, str(p))
                c.drawString(x + 55 * mm, y, str(cur))
                c.drawRightString(x + 112 * mm, y, f"{agg.get('in', 0):.2f}")
                c.drawRightString(x + 138 * mm, y, f"{agg.get('out', 0):.2f}")
                c.drawRightString(width - margin - 10 * mm, y, f"{agg.get('net', 0):.2f}")
                y -= 6 * mm

        c.showPage()
        c.save()
        pdf = buf.getvalue()
        buf.close()

        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="report_{period}.pdf"'
        return resp

    def _ensure_document_number(self, tx):
        if tx.document_number:
            return tx.document_number

        direction = tx.direction
        seq, _ = FinancialDocumentSequence.objects.get_or_create(direction=direction, defaults={'last_number': 0})
        seq.last_number += 1
        seq.save(update_fields=['last_number'])

        prefix = 'DOC-IN' if direction == 'in' else 'DOC-OUT'
        tx.document_number = f"{prefix}-{seq.last_number:06d}"
        tx.save(update_fields=['document_number'])
        return tx.document_number

    def _ensure_receipt_pdf(self, tx, request):
        if tx.receipt_pdf:
            return tx.receipt_pdf

        self._ensure_document_number(tx)
        pdf_bytes = self._build_receipt_pdf_bytes(tx, request)

        from django.core.files.base import ContentFile
        filename = f"receipt_{tx.document_number}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        tx.receipt_pdf.save(filename, ContentFile(pdf_bytes), save=False)
        tx.save(update_fields=['receipt_pdf'])
        return tx.receipt_pdf

    def _pdf_brand_layout(self, c, doc_title='', doc_subtitle='', badge_text='', badge_rgb=(0.12, 0.31, 0.47)):
        width, height = A4
        margin = 18 * mm

        # Branding area at top
        c.setFillColorRGB(0.98, 0.98, 0.99)
        c.rect(0, height - 35 * mm, width, 35 * mm, fill=1, stroke=0)

        # Logo placeholder
        logo_w = 28 * mm
        logo_h = 12 * mm
        c.setFillColorRGB(0.12, 0.31, 0.47)
        c.setFont('Helvetica-Bold', 10)
        c.drawString(margin, height - 12 * mm, 'Consolation et Paix Divine')

        # Badge
        if badge_text:
            c.setFillColorRGB(*badge_rgb)
            badge_w = max(20 * mm, c.stringWidth(badge_text, 'Helvetica-Bold', 9) + 6 * mm)
            c.roundRect(width - margin - badge_w, height - 20 * mm, badge_w, 8 * mm, 4, fill=1, stroke=0)
            c.setFillColorRGB(1, 1, 1)
            c.setFont('Helvetica-Bold', 9)
            c.drawCentredString(width - margin - badge_w / 2, height - 17.5 * mm, badge_text)

        # Title
        c.setFillColorRGB(0.07, 0.10, 0.16)
        c.setFont('Helvetica-Bold', 14)
        c.drawString(margin, height - 26 * mm, doc_title)

        if doc_subtitle:
            c.setFillColorRGB(0.40, 0.45, 0.55)
            c.setFont('Helvetica', 9)
            c.drawString(margin, height - 31 * mm, doc_subtitle)

        # Content start position
        card_top = height - 42 * mm
        y_start = card_top - 10 * mm

        return {
            'width': width,
            'height': height,
            'margin': margin,
            'x': margin,
            'y_start': y_start,
            'card_top': card_top,
        }

    def _build_receipt_pdf_bytes(self, tx, request):
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)

        meta = self._pdf_brand_layout(
            c,
            doc_title='Reçu de don',
            doc_subtitle='Document officiel — conservation recommandée',
            badge_text='REÇU',
            badge_rgb=(0.05, 0.65, 0.45),
        )

        width = meta['width']
        margin = meta['margin']
        x = meta['x']
        y = meta['y_start']

        c.setFillColorRGB(0.07, 0.10, 0.16)
        c.setFont('Helvetica-Bold', 16)
        c.drawString(x, y, f"{tx.amount:,.2f} {tx.currency or 'CDF'}")
        y -= 8 * mm

        c.setFillColorRGB(0.20, 0.24, 0.30)
        c.setFont('Helvetica', 10)
        c.drawString(x, y, f"Date: {tx.date.isoformat() if tx.date else '—'}")
        y -= 6 * mm
        c.drawString(x, y, f"Donateur: {tx.donor_name or '—'}")
        y -= 6 * mm
        c.drawString(x, y, f"Type: {tx.transaction_type or '—'}")
        y -= 6 * mm
        c.drawString(x, y, f"Référence: {tx.document_number or tx.receipt_code or '—'}")
        y -= 6 * mm
        if tx.description:
            c.drawString(x, y, f"Description: {tx.description[:100]}")

        # Footer
        c.setFillColorRGB(0.45, 0.50, 0.60)
        c.setFont('Helvetica', 8)
        c.drawRightString(width - margin, 10 * mm, f"Généré le {timezone.now().strftime('%Y-%m-%d %H:%M')}")

        c.showPage()
        c.save()
        return buf.getvalue()

    def _build_voucher_pdf_bytes(self, tx, request):
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)

        meta = self._pdf_brand_layout(
            c,
            doc_title='Bon de sortie',
            doc_subtitle='Document officiel — pièce justificative',
            badge_text='SORTIE',
            badge_rgb=(0.75, 0.25, 0.20),
        )

        width = meta['width']
        margin = meta['margin']
        x = meta['x']
        y = meta['y_start']

        c.setFillColorRGB(0.07, 0.10, 0.16)
        c.setFont('Helvetica-Bold', 16)
        c.drawString(x, y, f"{tx.amount:,.2f} {tx.currency or 'CDF'}")
        y -= 8 * mm

        c.setFillColorRGB(0.20, 0.24, 0.30)
        c.setFont('Helvetica', 10)
        c.drawString(x, y, f"Date: {tx.date.isoformat() if tx.date else '—'}")
        y -= 6 * mm
        c.drawString(x, y, f"Bénéficiaire: {tx.recipient_name or '—'}")
        y -= 6 * mm
        c.drawString(x, y, f"Type: {tx.transaction_type or '—'}")
        y -= 6 * mm
        c.drawString(x, y, f"Document: {tx.document_number or '—'}")

        # Footer
        c.setFillColorRGB(0.45, 0.50, 0.60)
        c.setFont('Helvetica', 8)
        c.drawRightString(width - margin, 10 * mm, f"Généré le {timezone.now().strftime('%Y-%m-%d %H:%M')}")

        c.showPage()
        c.save()
        return buf.getvalue()
